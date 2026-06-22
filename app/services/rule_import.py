"""XLSX rule import & merge — ports src/lib/ruleImport.ts.

One row = one compliance rule. Multi-value cells (clauses, actions, artifacts)
accept newline / semicolon / pipe separators. Persona & Trigger ids are derived
from labels when an explicit id column is omitted.
"""
from __future__ import annotations

import io
import re
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook

from app.schemas.common import ImportIssue
from app.schemas.simulator import ComplianceRule, Persona, TriggerEvent

TEMPLATE_COLUMNS = [
    "Law",
    "Law ID",
    "Persona",
    "Persona ID",
    "Trigger Label",
    "Trigger ID",
    "Rule ID",
    "Title",
    "Tier",
    "Clauses",
    "Actions",
    "Artifacts",
    "Penalty (crore)",
    "Penalty Basis",
    "Penalty Ref",
]

TEMPLATE_SAMPLE_ROW = {
    "Law": "Digital Personal Data Protection Act",
    "Law ID": "dpdp-act-2023",
    "Persona": "Significant Data Fiduciary",
    "Persona ID": "significant-data-fiduciary",
    "Trigger Label": "sdf_designation_received",
    "Trigger ID": "sdf_designation_received",
    "Rule ID": "DPDP-SDF-ECP-9",
    "Title": "Example: appoint a compliance officer",
    "Tier": "mandatory",
    "Clauses": "Section 10",
    "Actions": "Step one\nStep two",
    "Artifacts": "Appointment Record\nBoard Resolution",
    "Penalty (crore)": 150,
    "Penalty Basis": "Failure to comply as SDF",
    "Penalty Ref": "Schedule 1 Item 5",
}


def _slugify(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return re.sub(r"^-+|-+$", "", value)


def _split_list(value: Any) -> List[str]:
    if value is None:
        return []
    return [s.strip() for s in re.split(r"[\n;|]+", str(value)) if s.strip()]


def _str(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _pick(row: Dict[str, Any], *names: str) -> Any:
    """Case-insensitive column lookup (forgiving of header casing/spacing)."""
    lowered = {k.strip().lower(): v for k, v in row.items()}
    for name in names:
        if name.lower() in lowered:
            return lowered[name.lower()]
    return None


def _normalise_tier(value: Any) -> Optional[str]:
    v = _str(value).lower()
    if v in ("mandatory", "required", "m"):
        return "mandatory"
    if v in ("recommended", "optional", "r"):
        return "recommended"
    return None


def parse_rows(rows: List[Dict[str, Any]]) -> Tuple[List[Persona], int, List[ImportIssue]]:
    """Parse spreadsheet rows into the nested Persona -> Trigger -> Rule structure."""
    issues: List[ImportIssue] = []
    persona_map: "Dict[str, Persona]" = {}
    seen_rule_ids: set = set()
    rule_count = 0

    for i, row in enumerate(rows):
        row_num = i + 1

        law_label = _str(_pick(row, "Law", "Law Label"))
        persona_label = _str(_pick(row, "Persona", "Persona Label"))
        trigger_label = _str(_pick(row, "Trigger Label", "Trigger"))
        rule_id = _str(_pick(row, "Rule ID", "RuleID", "ID"))
        title = _str(_pick(row, "Title", "Rule Title"))
        tier = _normalise_tier(_pick(row, "Tier"))
        penalty_raw = _pick(row, "Penalty (crore)", "Penalty", "Penalty Crore")

        # Skip fully blank rows silently.
        if (
            not law_label
            and not persona_label
            and not trigger_label
            and not rule_id
            and not title
        ):
            continue

        row_issues: List[str] = []
        if not law_label:
            row_issues.append("missing Law")
        if not persona_label:
            row_issues.append("missing Persona")
        if not trigger_label:
            row_issues.append("missing Trigger Label")
        if not rule_id:
            row_issues.append("missing Rule ID")
        if not title:
            row_issues.append("missing Title")
        if not tier:
            row_issues.append("Tier must be 'mandatory' or 'recommended'")

        try:
            penalty = float(penalty_raw)
            penalty_ok = True
        except (TypeError, ValueError):
            penalty_ok = False
            penalty = 0.0

        if penalty_raw is None or penalty_raw == "" or not penalty_ok:
            row_issues.append("Penalty must be a number (in crore)")
        elif penalty < 0:
            row_issues.append("Penalty cannot be negative")

        if rule_id and rule_id in seen_rule_ids:
            row_issues.append(f'duplicate Rule ID "{rule_id}" in this file')

        if row_issues:
            issues.append(ImportIssue(row=row_num, message="; ".join(row_issues)))
            continue

        seen_rule_ids.add(rule_id)

        law_id = _str(_pick(row, "Law ID", "LawID")) or _slugify(law_label)
        persona_id = _str(_pick(row, "Persona ID", "PersonaID")) or _slugify(persona_label)
        trigger_id = _str(_pick(row, "Trigger ID", "TriggerID")) or _slugify(trigger_label)

        rule = ComplianceRule(
            id=rule_id,
            title=title,
            tier=tier,  # type: ignore[arg-type]
            clauses=_split_list(_pick(row, "Clauses", "Clause")),
            actions=_split_list(_pick(row, "Actions", "Required Actions")),
            artifacts=_split_list(_pick(row, "Artifacts", "Compliance Artifacts")),
            penalty=penalty,
            penaltyBasis=_str(_pick(row, "Penalty Basis", "Basis")),
            penaltyRef=_str(_pick(row, "Penalty Ref", "Reference", "Ref")),
        )

        # Key by (law, persona) so the same persona id can appear under more
        # than one law without the rows colliding.
        persona_key = f"{law_id}::{persona_id}"
        persona = persona_map.get(persona_key)
        if persona is None:
            persona = Persona(
                id=persona_id,
                label=persona_label,
                lawId=law_id,
                lawLabel=law_label,
                triggers=[],
            )
            persona_map[persona_key] = persona

        trigger = next((t for t in persona.triggers if t.id == trigger_id), None)
        if trigger is None:
            trigger = TriggerEvent(id=trigger_id, label=trigger_label, rules=[])
            persona.triggers.append(trigger)

        trigger.rules.append(rule)
        rule_count += 1

    return list(persona_map.values()), rule_count, issues


def merge_personas(
    existing: List[Persona], incoming: List[Persona]
) -> Tuple[List[Persona], int, int]:
    """Merge imported personas into the existing set.

    Returns (merged, added, updated).
    """
    # Deep clone so we never mutate the source data.
    merged: List[Persona] = [Persona(**p.model_dump()) for p in existing]
    added = 0
    updated = 0

    for in_p in incoming:
        # Match on (lawId, id) so personas are scoped to their law.
        p = next(
            (m for m in merged if m.id == in_p.id and m.lawId == in_p.lawId), None
        )
        if p is None:
            p = Persona(
                id=in_p.id,
                label=in_p.label,
                lawId=in_p.lawId,
                lawLabel=in_p.lawLabel,
                triggers=[],
            )
            merged.append(p)
        elif in_p.lawLabel:
            # Refresh law label on an existing persona if the import supplies one.
            p.lawLabel = in_p.lawLabel
        for in_t in in_p.triggers:
            t = next((m for m in p.triggers if m.id == in_t.id), None)
            if t is None:
                t = TriggerEvent(id=in_t.id, label=in_t.label, rules=[])
                p.triggers.append(t)
            for in_r in in_t.rules:
                idx = next((j for j, r in enumerate(t.rules) if r.id == in_r.id), -1)
                if idx >= 0:
                    t.rules[idx] = in_r
                    updated += 1
                else:
                    t.rules.append(in_r)
                    added += 1

    return merged, added, updated


# --- Spreadsheet IO -------------------------------------------------------

def workbook_to_rows(content: bytes) -> List[Dict[str, Any]]:
    """Read the first sheet of an .xlsx workbook into a list of dict rows."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    headers = [str(h).strip() if h is not None else "" for h in header]
    out: List[Dict[str, Any]] = []
    for raw in rows_iter:
        row = {headers[i]: (raw[i] if i < len(raw) else None) for i in range(len(headers))}
        out.append(row)
    return out


def build_template_workbook() -> bytes:
    """Build the downloadable .xlsx template with one sample row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rules"
    ws.append(TEMPLATE_COLUMNS)
    ws.append([TEMPLATE_SAMPLE_ROW[c] for c in TEMPLATE_COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
